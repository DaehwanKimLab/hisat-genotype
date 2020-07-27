#!/usr/bin/env python
# --------------------------------------------------------------------------- #
# Copyright 2017, Daehwan Kim <infphilo@gmail.com>                            #
#                                                                             #
# This file is part of HISAT-genotype. This converts codis data to a format   #
# for use in HISAT-genotype                                                   #
#                                                                             #
# HISAT-genotype is free software: you can redistribute it and/or modify      #
# it under the terms of the GNU General Public License as published by        #
# the Free Software Foundation, either version 3 of the License, or           #
# (at your option) any later version.                                         #
#                                                                             #
# HISAT-genotype is distributed in the hope that it will be useful,           #
# but WITHOUT ANY WARRANTY; without even the implied warranty of              #
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the               #
# GNU General Public License for more details.                                #
#                                                                             #
# You should have received a copy of the GNU General Public License           #
# along with HISAT-genotype.  If not, see <http://www.gnu.org/licenses/>.     #
# --------------------------------------------------------------------------- #

import os
import sys
import subprocess
import re
import inspect, operator
from copy import deepcopy
from argparse import ArgumentParser, FileType
import hisatgenotype_typing_common as typing_common
import hisatgenotype_args as arguments
try:
    import openpyxl
except ImportError:
    print("Error: please install openpyxl by running 'pip install openpyxl'.", file=sys.stderr)
    sys.exit(1)

# --------------------------------------------------------------------------- #
# Global settings                                                             #
# --------------------------------------------------------------------------- #
""" 
sequences for DNA fingerprinting loci are available at 
http://www.cstl.nist.gov/biotech/strbase/seq_ref.htm
"""
orig_CODIS_seq = {
    "CSF1PO" :
    # http://www.cstl.nist.gov/biotech/strbase/str_CSF1PO.htm
    # allele 13: 5:150076172-150076490 - (samtools faidx genome.fa - GRCh38)
    ["[AGAT]13",
     "AACCTGAGTCTGCCAAGGACTAGCAGGTTGCTAACCACCCTGTGTCTCAGTTTTCCTACCTGTAAAA"\
         "TGAAGATATTAACAGTAACTGCCTTCATAGATAGAAGATAGATAGATT", # left flanking sequence
     "AGATAGATAGATAGATAGATAGATAGATAGATAGATAGATAGATAGATAGAT", # STR
     "AGGAAGTACTTAGAACAGGGTCTGACACAGGAAATGCTGTCCAAGTGTGCACCAGGAGATAGTATCTG"\
         "AGAAGGCTCAGTCTGGCACCATGTGGGTTGGGTGGGAACCTGGAGGCTGGAGAATGGGCTGAAGA"\
         "TGGCCAGTGGTGTGTGGAA"], # right flanking sequence
             
    "FGA" :
    # http://www.cstl.nist.gov/biotech/strbase/str_FGA.htm
    # allele 22: 4:154587696-154587891 -
    ["[TTTC]3TTTTTTCT[CTTT]14CTCC[TTCC]2",
     "GCCCCATAGGTTTTGAACTCACAGATTAAACTGTAACCAAAATAAAATTAGGCATATTTACAAGCTAG",
     "TTTCTTTCTTTCTTTTTTCTCTTTCTTTCTTTCTTTCTTTCTTTCTTTCTTTCTTTCTTTCTTTCTTTCT"\
         "TTCTTTCTCCTTCCTTCC",
     "TTTCTTCCTTTCTTTTTTGCTGGCAATTACAGACAAATCA"],

    "TH01" :
    # http://www.cstl.nist.gov/biotech/strbase/str_TH01.htm
    # allele 7: 11:2170990-2171176 +
    ["[AATG]7",
     "GTGGGCTGAAAAGCTCCCGATTATCCAGCCTGGCCCACACAGTCCCCTGTACACAGGGCTTCCGAGTGCAG"\
         "GTCACAGGGAACACAGACTCCATGGTG",
     "AATGAATGAATGAATGAATGAATGAATG",
     "AGGGAAATAAGGGAGGAACAGGCCAATGGGAATCACCCCAGAGCCCAGATACCCTTTGAAT"],
             
    "TPOX" :
    # http://www.cstl.nist.gov/biotech/strbase/str_TPOX.htm
    # allele 8: 2:1489617-1489848
    ["[AATG]8",
     "ACTGGCACAGAACAGGCACTTAGGGAACCCTCACTG",
     "AATGAATGAATGAATGAATGAATGAATGAATG",
     "TTTGGGCAAATAAACGCTGACAAGGACAGAAGGGCCTAGCGGGAAGGGAACAGGAGTAAGACCAGCGCACAGC"\
         "CCGACTTGTGTTCAGAAGACCTGGGATTGGACCTGAGGAGTTCAATTTTGGATGAATCTCTTAATTAACC"\
         "TGTGGGGTTCCCAGTTCCTCC"],
             
    "VWA" :
    # http://www.cstl.nist.gov/biotech/strbase/str_VWA.htm
    # allele unknown: 12:5983938-5984087 -
    ["TCTA[TCTG]5[TCTA]11TCCA TCTA",
     "CCCTAGTGGATGATAAGAATAATCAGTATGTGACTTGGATTGA",
     "TCTATCTGTCTGTCTGTCTGTCTGTCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCCATCTA",
     "TCCATCCATCCTATGTATTTATCATCTGTCC"],
             
    "D3S1358" :
    # http://www.cstl.nist.gov/biotech/strbase/str_D3S1358.htm
    # allele unknown: 3:45540713-45540843 +
    ["TCTATCTG[TCTA]14",
     "ATGAAATCAACAGAGGCTTGCATGTA",
     "TCTATCTGTCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTA",
     "TGAGACAGGGTCTTGCTCTGTCACCCAGATTGGACTGCAGT"],
             
    "D5S818" :
    # http://www.cstl.nist.gov/biotech/strbase/str_D5S818.htm
    # allele 11: 5:123775504-123775638 -
    ["[AGAT]11",
     "GGTGATTTTCCTCTTTGGTATCCTTATGTAATATTTTGA",
     "AGATAGATAGATAGATAGATAGATAGATAGATAGATAGATAGAT",
     "AGAGGTATAAATAAGGATACAGATAAAGATACAAATGTTGTAAACTGTGGCT"],
             
    "D7S820" :
    # http://www.cstl.nist.gov/biotech/strbase/str_D7S820.htm
    # allele 13: 7:84160125-84160367 -
    ["[GATA]13",
     "ATGTTGGTCAGGCTGACTATGGAGTTATTTTAAGGTTAATATATATAAAGGGTATGATAGAACACTTGTCATA"\
         "GTTTAGAACGAACTAAC",
     "GATAGATAGATAGATAGATAGATAGATAGATAGATAGATAGATAGATAGATA",
     "GACAGATTGATAGTTTTTTTTAATCTCACTAAATAGTCTATAGTAAACATTTAATTACCAATATTTGGTGCAAT"\
         "TCTGTCAATGAGGATAAATGTGGAATC"],
             
    "D8S1179" :
    # http://www.cstl.nist.gov/biotech/strbase/str_D8S1179.htm
    # allele 13: 8:124894838-124895018 +
    ["[TCTA]1[TCTG]1[TCTA]11",
     "TTTTTGTATTTCATGTGTACATTCGTA",
     "TCTATCTGTCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTA",
     "TTCCCCACAGTGAAAATAATCTACAGGATAGGTAAATAAATTAAGGCATATTCACGCAATGGGATACGATAC"\
         "AGTGATGAAAATGAACTAATTATAGCTACG"],
             
    "D13S317" :
    # http://www.cstl.nist.gov/biotech/strbase/str_D13S317.htm
    # Perhaps, allele 11: 13:82147921-82148112 +
    ["[TATC]11A",
     "ATCACAGAAGTCTGGGATGTGGAGGAGAGTTCATTTCTTTAGTGGGCATCCGTGACTCTCTGGACTCTGACC"\
         "CATCTAACGCCTATCTGTATTTACAAATACAT",
     "TATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCA",
     "ATCAATCATCTATCTATCTTTCTGTCTGTCTTTTTGGGCTGCC"],
             
    "D16S539" :
    # http://www.cstl.nist.gov/biotech/strbase/str_D16S539.htm
    # allele 11: 16:86352518-86352805 +
    ["[GATA]11",
     "GGGGGTCTAAGAGCTTGTAAAAAGTGTACAAGTGCCAGATGCTCGTTGTGCACAAATCTAAATGCAGAAAAGC"\
         "ACTGAAAGAAGAATCCAGAAAACCACAGTTCCCATTTTTATATGGGAGCAAACAAAGGCAGATCCCAAG"\
         "CTCTTCCTCTTCCCTAGATCAATACAGACAGACAGACAGGTG",
     "GATAGATAGATAGATAGATAGATAGATAGATAGATAGATAGATA",
     "TCATTGAAAGACAAAACAGAGATGGATGATAGATACATGCTTACAGATGCACACACAAAC"],
             
    "D18S51" :
    # http://www.cstl.nist.gov/biotech/strbase/str_D18S51.htm
    # allele 18: 18:63281611-63281916 +
    ["[AGAA]18",
     "GAGCCATGTTCATGCCACTGCACTTCACTCTGAGTGACAAATTGAGACCTTGTCTC",
     "AGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAAAGAA",
     "AAAGAGAGAGGAAAGAAAGAGAAAAAGAAAAGAAATAGTAGCAACTGTTATTGTAAGACATCTCCACACACCAG"\
         "AGAAGTTAATTTTAATTTTAACATGTTAAGAACAGAGAGAAGCCAACATGTCCACCTTAGGCTGACGGTTT"\
         "GTTTATTTGTGTTGTTGCTGGTAGTCGGGTTTG"],
             
    "D21S11" :
    # http://www.cstl.nist.gov/biotech/strbase/str_D21S11.htm
    # Perhaps, allele 29: 21:19181945-19182165 +
    ["[TCTA]4[TCTG]6[TCTA]3TA[TCTA]3TCA[TCTA]2TCCATA[TCTA]11",
     "GTGAGTCAATTCCCCAAGTGAATTGCCT",
     "TCTATCTATCTATCTATCTGTCTGTCTGTCTGTCTGTCTGTCTATCTATCTATATCTATCTATCTATCATCTATCT"\
         "ATCCATATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTATCTA",
     "TCGTCTATCTATCCAGTCTATCTACCTCCTATTAGTCTGTCTCTGGAGAACATTGACTAATACAAC"],

    # "AMEL" - http://www.cstl.nist.gov/biotech/strbase/jpg_amel.htm
    #          X chromosome has 6 bp deletion and Y chromosome doesn't
    "AMELX" :
    ["",
     "TGTTGATTCTTTATCCCAGATGTTTCTCAAGTGG", # chromosome X at 11296898
     "",
     ""],

    "AMELY" :
    ["",
     "AGAAACCACTTTATTTGGGATGAAGAATCCACC", # chromosome Y at 6869902
     "",
     ""]
}

CODIS_ref_name = {}


# --------------------------------------------------------------------------- #
# Functions for handling sequences                                            #
# --------------------------------------------------------------------------- #
def get_flanking_seqs(seq,
                      flank_len = 500):
    def align_seq(seq):
        aligner_cmd = ["hisat2",
                       "--score-min", "C,0",
                       "--no-unal",
                        "-x", "grch38/genome",
                        "-c", seq]
        align_proc = subprocess.Popen(aligner_cmd,
                                      universal_newlines=True,
                                      stdout=subprocess.PIPE,
                                      stderr=open("/dev/null", 'w'))
        chr, left, right, strand = "", -1, -1, '+'
        for line in align_proc.stdout:
            if line.startswith('@'):
                continue
            line = line.strip()
            cols = line.split()
            allele_id, flag, chr, left, _, cigar_str = cols[:6]
            assert cigar_str[-1] == 'M'
            left = int(left)
            flag = int(flag)
            strand = '-' if flag & 0x10 else '+'
            assert cigar_str == ("%dM" % len(seq))
            right = left + len(seq)
            break
        
        assert chr != "" and left >= 0 and right > left
        return chr, left, right, strand
    
    chr, left, right, strand = align_seq(seq)    
    left_flank_seq  = ""
    right_flank_seq = ""
    if left > 1:
        extract_seq_cmd = ["samtools", 
                           "faidx", 
                           "genome.fa", 
                           "%s:%d-%d" % (chr, 
                                         max(1, left - flank_len), 
                                         left - 1)]
        extract_seq_proc = subprocess.Popen(extract_seq_cmd,
                                            universal_newlines = True,
                                            stdout=subprocess.PIPE,
                                            stderr=open("/dev/null", 'w'))
        for line in extract_seq_proc.stdout:
            if line.startswith('>'):
                continue
            line = line.strip()
            left_flank_seq += line
    extract_seq_cmd = ["samtools", 
                       "faidx", 
                       "genome.fa", 
                       "%s:%d-%d" % (chr, right, right + flank_len - 1)]
    extract_seq_proc = subprocess.Popen(extract_seq_cmd,
                                        universal_newlines=True,
                                        stdout=subprocess.PIPE,
                                        stderr=open("/dev/null", 'w'))
    for line in extract_seq_proc.stdout:
        if line.startswith('>'):
            continue
        line = line.strip()
        right_flank_seq += line

    if strand == '-':
        left_flank_seq  = typing_common.reverse_complement(right_flank_seq)
        right_flank_seq = typing_common.reverse_complement(left_flank_seq)

    chr, _, _, _ = align_seq(left_flank_seq + seq + right_flank_seq)
    assert chr != ""
    
    return left_flank_seq, right_flank_seq

def get_equal_score(repeat_i, 
                    repeat_nums_i, 
                    repeat_j, 
                    repeat_nums_j):
    if repeat_i == repeat_j:
        # DK - experimental SW alignment
        min_diff = sys.maxsize
        for repeat_num_i in repeat_nums_i:
            for repeat_num_j in repeat_nums_j:
                min_diff = min(abs(repeat_num_i - repeat_num_j), min_diff)
        equal_score = -min_diff / 10.0\
                         + (len(repeat_nums_i) \
                         + len(repeat_nums_j)) / 100.0
        equal_score = max(min(0.0 if min_diff == 0 else -0.1, equal_score), -0.9)

        # DK - just for now
        equal_score = 0
        
        return equal_score
    elif repeat_nums_i == repeat_nums_j and repeat_nums_i == set([1]):
        return -1
    else:
        return -2

""" Smith Waterman Algorithm """
def SW_alignment(allele_i, allele_j):
    n, m = len(allele_i), len(allele_j)
    a = [[-(i+j) if i == 0 or j == 0 \
                 else 0 for j in range(m + 1)] for i in range(n + 1)]

    # Fill 2D array
    for i in range(n):
        repeat_i, repeat_nums_i = allele_i[i]
        for j in range(m):
            repeat_j, repeat_nums_j = allele_j[j]
            equal_score = get_equal_score(repeat_i, 
                                          repeat_nums_i, 
                                          repeat_j, 
                                          repeat_nums_j)
            a[i+1][j+1] = max(a[i][j+1] - 1, a[i+1][j] - 1, a[i][j] + equal_score)

    return a, n, m

def combine_alleles(backbone_allele, add_allele):
    allele_i = backbone_allele
    allele_j = add_allele
    a, n, m = SW_alignment(allele_i, allele_j)

    # Back tracking
    new_backbone_allele = []
    i = n - 1
    j = m - 1
    while i >= 0 or j >= 0:
        if i < 0:
            repeat_j, repeat_nums_j = allele_j[j]
            new_backbone_allele.append([repeat_j, repeat_nums_j | set([0])])
            j -= 1
        elif j < 0:
            repeat_i, repeat_nums_i = allele_i[i]
            new_backbone_allele.append([repeat_i, repeat_nums_i | set([0])])
            i -= 1
        else:
            repeat_i, repeat_nums_i = allele_i[i]
            repeat_j, repeat_nums_j = allele_j[j]    
            equal_score = get_equal_score(repeat_i, 
                                          repeat_nums_i, 
                                          repeat_j, 
                                          repeat_nums_j)
            if a[i][j+1] - 1 == a[i+1][j+1]:
                new_backbone_allele.append([repeat_i, repeat_nums_i | set([0])])
                i -= 1
            elif a[i+1][j] - 1 == a[i+1][j+1]:
                new_backbone_allele.append([repeat_j, repeat_nums_j | set([0])])
                j -= 1
            else:
                assert a[i][j] + equal_score == a[i+1][j+1]
                if repeat_i == repeat_j:
                    new_backbone_allele.append([repeat_i, 
                                                repeat_nums_i | repeat_nums_j])
                else:
                    assert repeat_nums_i == repeat_nums_j
                    assert repeat_nums_i == set([1])
                    new_backbone_allele.append([repeat_i | repeat_j, 
                                                repeat_nums_i | repeat_nums_j])
                i -= 1
                j -= 1

    new_backbone_allele = new_backbone_allele[::-1]
    return new_backbone_allele

def msf_alignment(backbone_allele, allele):
    allele_i = backbone_allele
    allele_j = allele
    a, n, m = SW_alignment(allele_i, allele_j)

    # Back tracking
    allele_seq   = ""
    backbone_seq = ""
    i = n - 1
    j = m - 1
    while i >= 0 or j >= 0:
        assert i >= 0
        repeats_i, repeat_nums_i = allele_i[i]
        repeat_i   = ""
        max_repeat = ""
        for repeat_str in repeats_i:
            if len(repeat_str) > len(repeat_i):
                repeat_i = repeat_str
        repeat_num_i = max(repeat_nums_i)
        if j < 0:
            allele_seq   = '.' * (len(repeat_i) * repeat_num_i) + allele_seq
            backbone_seq = repeat_i * repeat_num_i + backbone_seq
            i -= 1
        else:
            repeats_j, repeat_nums_j = allele_j[j]
            assert len(repeats_j) == 1 and len(repeat_nums_j) == 1
            repeat_j, repeat_num_j = list(repeats_j)[0], list(repeat_nums_j)[0]
            equal_score = get_equal_score(repeats_i, 
                                          repeat_nums_i, 
                                          repeats_j, 
                                          repeat_nums_j)
            if a[i][j+1] - 1 == a[i+1][j+1]:
                allele_seq   = '.' * (len(repeat_i) * repeat_num_i) + allele_seq
                backbone_seq = repeat_i * repeat_num_i + backbone_seq
                i -= 1
            else:
                assert a[i][j] + equal_score == a[i+1][j+1]
                if repeat_i == repeat_j:
                    add_seq      = repeat_i * repeat_num_j
                    dot_seq      = '.' \
                                    * (len(repeat_i) * (repeat_num_i - repeat_num_j))
                    allele_seq   = add_seq + dot_seq + allele_seq
                    add_seq      = repeat_i * repeat_num_i
                    backbone_seq = add_seq + backbone_seq                    
                else:
                    assert repeat_nums_i == repeat_nums_j and repeat_nums_i == set([1])
                    dot_seq      = '.' * (len(repeat_i) - len(repeat_j))
                    allele_seq   = repeat_j + dot_seq + allele_seq
                    backbone_seq = repeat_i + backbone_seq                    
                i -= 1
                j -= 1

    return allele_seq, backbone_seq

""" Extract multiple sequence alignments """
def extract_msa(base_dname,
                base_fname,
                locus_list,
                min_freq,
                verbose):    
    # Download human genome and HISAT2 index
    typing_common.download_genome_and_index()

    # Load allele frequency information
    allele_freq = {}
    if min_freq > 0.0:
        excel = openpyxl.load_workbook("hisatgenotype_db/CODIS/"\
                                         "NIST-US1036-AlleleFrequencies.xlsx")
        sheet = excel.get_sheet_by_name(u'All data, n=1036')
        for col in range(2, 100):
            locus_name = sheet.cell(row = 3, column = col).value
            if not locus_name:
                break
            locus_name = locus_name.encode('ascii','ignore')
            locus_name = locus_name.upper()
            assert locus_name not in allele_freq
            allele_freq[locus_name] = {}

            for row in range(4, 101):
                allele_id = sheet.cell(row = row, column = 1).value
                allele_id = str(allele_id)
                freq = sheet.cell(row = row, column = col).value
                if not freq:
                    continue
                allele_freq[locus_name][allele_id] = float(freq)
        excel.close()

    CODIS_seq = orig_CODIS_seq
    if len(locus_list) > 0:
        new_CODIS_seq = {}
        for locus_name, fields in CODIS_seq.items():
            if locus_name in locus_list:
                new_CODIS_seq[locus_name] = fields
        CODIS_seq = new_CODIS_seq        

    """ 
    Add some additional sequences to allele sequences to make
    them reasonably long for typing and assembly 
    """
    for locus_name, fields in CODIS_seq.items():
        _, left_seq, repeat_seq, right_seq = fields
        allele_seq = left_seq + repeat_seq + right_seq
        left_flank_seq, right_flank_seq = get_flanking_seqs(allele_seq)
        CODIS_seq[locus_name][1] = left_flank_seq + left_seq
        CODIS_seq[locus_name][3] = right_seq + right_flank_seq

        print( "%s is found on the reference genome (GRCh38)" % locus_name, 
              file=sys.stderr)
    
    for locus_name in CODIS_seq.keys():
        alleles = []
        for line in open("hisatgenotype_db/CODIS/codis.dat"):
            locus_name2, allele_id, repeat_st = line.strip().split('\t')
            if locus_name != locus_name2:
                continue
            if min_freq > 0.0:
                assert locus_name in allele_freq
                if allele_id not in allele_freq[locus_name] or \
                   allele_freq[locus_name][allele_id] < min_freq:
                    continue
                
            alleles.append([allele_id, repeat_st])

        # From   [TTTC]3TTTTTTCT[CTTT]20CTCC[TTCC]2
        # To     [['TTTC', [3]], ['TTTTTTCT', [1]], 
        #         ['CTTT', [20]], ['CTCC', [1]], ['TTCC', [2]]]
        def read_allele(repeat_st):
            allele = []
            s      = 0
            while s < len(repeat_st):
                ch = repeat_st[s]
                if ch == ' ':
                    s += 1
                    continue
                assert ch in "[ACGT"
                if ch == '[':
                    s     += 1
                    repeat = ""
                    while s < len(repeat_st):
                        nt = repeat_st[s]
                        if nt in "ACGT":
                            repeat += nt
                            s += 1
                        else:
                            assert nt == ']'
                            s += 1
                            break
                    assert s < len(repeat_st)
                    num = 0
                    while s < len(repeat_st):
                        digit = repeat_st[s]
                        if digit.isdigit():
                            num = num * 10 + int(digit)
                            s  += 1
                        else:
                            break
                    assert num > 0
                    allele.append([set([repeat]), set([num])])
                else:
                    repeat = ""
                    while s < len(repeat_st):
                        nt = repeat_st[s]
                        if nt in "ACGT":
                            repeat += nt
                            s += 1
                        else:
                            assert nt == ' ' or nt == '['
                            break
                    allele.append([set([repeat]), set([1])])

            # Sanity check
            cmp_repeat_st = ""
            for repeats, repeat_nums in allele:
                repeat = list(repeats)[0]
                repeat_num = list(repeat_nums)[0]
                if repeat_num > 1 or locus_name == "D8S1179":
                    cmp_repeat_st += "["
                cmp_repeat_st += repeat
                if repeat_num > 1 or locus_name == "D8S1179":
                    cmp_repeat_st += "]%d" % repeat_num

            assert repeat_st.replace(' ', '') == cmp_repeat_st.replace(' ', '')
            return allele

        alleles = [[allele_id, read_allele(repeat_st)] for allele_id, repeat_st in alleles]

        def to_sequence(repeat_st):
            sequence = ""
            for repeats, repeat_nums in repeat_st:
                repeat     = list(repeats)[0]
                repeat_num = list(repeat_nums)[0]
                sequence  += (repeat * repeat_num)
            return sequence

        def remove_redundant_alleles(alleles):
            seq_to_ids  = {}
            new_alleles = []
            for allele_id, repeat_st in alleles:
                allele_seq = to_sequence(repeat_st)
                if allele_seq in seq_to_ids:
                    print("Warning: %s: %s has the same sequence as %s" \
                            % (locus_name, allele_id, seq_to_ids[allele_seq]), 
                          file = sys.stderr)
                    continue
                if allele_seq not in seq_to_ids:
                    seq_to_ids[allele_seq] = [allele_id]
                else:
                    seq_to_ids[allele_seq].append(allele_id)         
                new_alleles.append([allele_id, repeat_st])

            return new_alleles

        alleles = remove_redundant_alleles(alleles)

        allele_seqs = [[allele_id, to_sequence(repeat_st)] \
                        for allele_id, repeat_st in alleles]

        ref_allele_st, \
          ref_allele_left, \
          ref_allele, \
          ref_allele_right \
            = CODIS_seq[locus_name]
        ref_allele_st = read_allele(ref_allele_st)
        for allele_id, allele_seq in allele_seqs:
            if ref_allele == allele_seq:
                CODIS_ref_name[locus_name] = allele_id
                break
            
        # Add GRCh38 allele
        if locus_name not in CODIS_ref_name:
            allele_id = "GRCh38"
            CODIS_ref_name[locus_name] = allele_id
            allele_seqs = [[allele_id, ref_allele]] + allele_seqs
            alleles = [[allele_id, ref_allele_st]] + alleles

        print("%s: %d alleles with reference allele as %s" \
                % (locus_name, len(alleles), CODIS_ref_name[locus_name]), 
              file=sys.stderr)
        if verbose:
            print(("\t", ref_allele_left, ref_allele, ref_allele_right), 
                  file=sys.stderr)
            for allele_id, allele in alleles:
                print((allele_id, "\t", allele), file=sys.stderr)

        # Create a backbone sequence
        assert len(alleles) > 0
        backbone_allele = deepcopy(alleles[-1][1])
        for allele_id, allele_st in reversed(alleles[:-1]):
            if verbose:
                print(file=sys.stderr)
                print(allele_id, 
                      file=sys.stderr)
                print(("backbone         :", backbone_allele), 
                      file=sys.stderr)
                print(("allele           :", allele_st), 
                      file=sys.stderr)
            backbone_allele = combine_alleles(backbone_allele, allele_st)
            msf_allele_seq, msf_backbone_seq = msf_alignment(backbone_allele, 
                                                             allele_st)
            if verbose:                
                print(("combined backbone:", backbone_allele), 
                      file=sys.stderr)
                print(("msf_allele_seq  :", msf_allele_seq), 
                      file=sys.stderr)
                print(("msf_backbone_seq:", msf_backbone_seq), 
                      file=sys.stderr)
                print(file=sys.stderr)

        allele_dic = {}
        for allele_id, allele_seq in allele_seqs:
            allele_dic[allele_id] = allele_seq

        allele_repeat_msf = {}
        for allele_id, allele_st in alleles:
            msf_allele_seq, msf_backbone_seq = msf_alignment(backbone_allele, 
                                                             allele_st)
            allele_repeat_msf[allele_id] = msf_allele_seq

        # Sanity check
        assert len(allele_dic) == len(allele_repeat_msf)
        repeat_len = None
        for allele_id, repeat_msf in allele_repeat_msf.items():
            if not repeat_len:
                repeat_len = len(repeat_msf)
            else:
                assert repeat_len == len(repeat_msf)

        # Creat full multiple sequence alignment
        ref_allele_id = CODIS_ref_name[locus_name]
        allele_msf = {}
        for allele_id, repeat_msf in allele_repeat_msf.items():
            allele_msf[allele_id] = ref_allele_left \
                                        + repeat_msf \
                                        + ref_allele_right

        # Make sure the length of allele ID is short, less than 20 characters
        max_allele_id_len = max([len(allele_id) \
                                  for allele_id in allele_dic.keys()])
        assert max_allele_id_len < 20

        # Write MSF (multiple sequence alignment file)
        msf_len   = len(ref_allele_left) + len(ref_allele_right) + repeat_len
        msf_fname = "%s.msf" % locus_name
        msf_file  = open(msf_fname, 'w')
        for s in range(0, msf_len, 50):
            for allele_id, msf in allele_msf.items():
                assert len(msf) == msf_len
                allele_name = "%s*%s" % (locus_name, allele_id)
                print("%20s" % allele_name, file=msf_file)
                for s2 in range(s, min(msf_len, s + 50), 10):
                    print(" %s" % msf[s2:s2+10], 
                          file=msf_file)
                print(file=msf_file)

            if s + 50 >= msf_len:
                break
            print(file=msf_file)
        msf_file.close()

        # Write FASTA file
        fasta_fname = "%s_gen.fasta" % locus_name
        fasta_file  = open(fasta_fname, 'w')
        for allele_id, allele_seq in allele_seqs:
            gen_seq = ref_allele_left + allele_seq + ref_allele_right
            print(">%s*%s %d bp" % (locus_name, allele_id, len(gen_seq)), 
                  file=fasta_file)
            for s in range(0, len(gen_seq), 60):
                print(gen_seq[s:s+60], file=fasta_file)
        fasta_file.close()


# --------------------------------------------------------------------------- #
# Main function                                                               #
# --------------------------------------------------------------------------- #
if __name__ == '__main__':
    parser = ArgumentParser(
        description="Extract multiple sequence alignments "\
                        "for DNA Fingerprinting loci")
    
    # Add Arguments
    arguments.args_databases(parser)
    arguments.args_convert_codis(parser)
    arguments.args_common(parser,
                          threads=False) # No threading option

    args = parser.parse_args()
    if not base_fname:
        base_fname = 'codis'
        base_dname = ''
    elif args.base_fname.find('/') != -1:
        elems = args.base_fname.split('/')
        base_fname = elems[-1]
        base_dname = '/'.join(elems[:-1])
    else:
        base_fname = args.base_fname
        base_dname = ""
    if args.locus_list != "":
        locus_list = args.locus_list.split(',')
    else:
        locus_list = []
        
    extract_msa(base_dname,
                base_fname,
                locus_list,
                args.min_freq,
                args.verbose)

